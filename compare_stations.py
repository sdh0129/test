import os
import sys
import argparse
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import numpy as np
import matplotlib.pyplot as plt

# 기존 스크립트 모듈 가져오기
from io_excel import read_input
from algorithm import EnhancedASA

# Matplotlib 기본 폰트 설정 (한글 인코딩 문제를 피하기 위해 영문 표기를 기본으로 함)
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans', 'Calibri']


def run_comparison(input_filepath, mode='balanced', seed=42):
    """
    10개부터 20개 스테이션까지 최적화 알고리즘을 각각 실행하고 지표를 수집합니다.
    """
    print(f"\n{'='*70}")
    print(f"  SP3 CTR FLR 스테이션 수(10~20개)별 최적화 비교 분석 가동")
    print(f"{'='*70}\n")

    # 입력 엑셀 데이터 로드
    data = read_input(input_filepath)
    tasks = data['tasks']
    edges = data['edges']

    # 실행 모드 설정
    if mode == 'quick':
        params = {'T0': 50, 'Tf': 1.0, 'cr': 0.97, 'maxIT': 30}
        mode_label = "[Quick] Quick 모드"
    elif mode == 'balanced':
        params = {'T0': 80, 'Tf': 0.5, 'cr': 0.985, 'maxIT': 50}
        mode_label = "[Balanced] Balanced 모드"
    else:  # standard
        params = {'T0': 100, 'Tf': 0.1, 'cr': 0.99, 'maxIT': 100}
        mode_label = "[Standard] Standard 표준 모드"

    print(f"[1/3] 데이터 로드 완료. (총 공정 수: {len(tasks)}개, 선후관계: {len(edges)}쌍)")
    print(f"      가동 모드: {mode_label}\n")

    results = []

    # 10부터 20까지 스테이션 수 루프 실행
    for ns in range(15, 26):
        print(f"      [Run] Station {ns:02d} 최적화 진행 중... ", end='', flush=True)
        t_start = time.time()

        # 시드 설정 (동등한 조건 비교를 위해 매 실행 전 고정)
        np.random.seed(seed)
        import random
        random.seed(seed)

        algo = EnhancedASA(
            tasks=tasks,
            edges=edges,
            ns=ns,
            nh=1,  # 스테이션당 인간 작업자 수 상한
            nr=1,  # 스테이션당 로봇 수 상한
            **params
        )

        try:
            best = algo.run(verbose=False)
            elapsed = time.time() - t_start
            print(f"완료 ({elapsed:.1f}초) | CT: {best['cycle_time']:.1f}s | 라인 효율: {best['line_efficiency']*100:.1f}%")

            results.append({
                'ns': ns,
                'best_result': best,
                'cycle_time': best['cycle_time'],
                'efficiency': best['line_efficiency'],
                'smoothness': best['smoothness'],
                'humans': best['n_humans'],
                'robots': best['n_robots'],
                'workers': best['n_humans'] + best['n_robots'],
                'bottleneck_station': best['bottleneck_station'],
                'bottleneck_ct': best['bottleneck_ct'],
            })
        except Exception as e:
            print(f"실패 (에러: {e})")

    return results, tasks


def write_comparison_excel(results, tasks, output_dir):
    """
    수집된 비교 지표를 모아 스타일링이 가미된 엑셀 보고서로 저장하고,
    각 스테이션 수별 상세 일정을 시트별로 추가합니다.
    """
    report_path = os.path.join(output_dir, "comparison_report.xlsx")
    wb = openpyxl.Workbook()

    # ----------------------------------------------------
    # 1. Summary 시트 작성
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Comparison_Summary"
    
    # A1 타이틀 설정
    ws_summary['A1'] = "SP3 CTR FLR 조립 라인 밸런싱 스테이션별 비교 분석 결과"
    ws_summary['A1'].font = Font(name='Malgun Gothic', size=16, bold=True, color='1F4E79')
    ws_summary.row_dimensions[1].height = 30

    # 테이블 헤더 스타일 설정
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_font = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFF')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    
    headers = [
        "스테이션 수 (NS)", 
        "사이클 타임 (CT, 초)", 
        "라인 밸런스 효율 (%)", 
        "Smoothness Index", 
        "사용 인간 수 (NH)", 
        "사용 로봇 수 (NR)", 
        "총 투입 자원 수 (NH+NR)", 
        "병목 스테이션", 
        "병목 부하 (초)"
    ]

    for c_idx, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=c_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws_summary.row_dimensions[3].height = 25

    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 데이터 작성
    for r_idx, res in enumerate(results, 4):
        row_values = [
            res['ns'],
            res['cycle_time'],
            res['efficiency'],
            res['smoothness'],
            res['humans'],
            res['robots'],
            res['workers'],
            f"St-{res['bottleneck_station']:02d}",
            res['bottleneck_ct']
        ]

        for c_idx, val in enumerate(row_values, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.border = thin_border
            
            # 셀 가독성을 위한 열별 포맷 및 정렬
            if c_idx in [1, 5, 6, 7, 8]:
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')

            if c_idx in [2, 4, 9]:
                cell.number_format = '0.0'
            elif c_idx == 3:
                cell.number_format = '0.0%'
            elif c_idx in [1, 5, 6, 7]:
                cell.number_format = '#,##0'

        ws_summary.row_dimensions[r_idx].height = 20

    # 컬럼 너비 자동 맞춤
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ----------------------------------------------------
    # 2. 각 스테이션 수별 상세 스케줄 시트 작성
    # ----------------------------------------------------
    for res in results:
        ns = res['ns']
        sheet_name = f"St-{ns:02d}"
        ws_det = wb.create_sheet(title=sheet_name)

        # 소제목 작성
        ws_det['A1'] = f"상세 할당 정보 (스테이션 수: {ns}개 배치 결과)"
        ws_det['A1'].font = Font(name='Malgun Gothic', size=14, bold=True, color='305496')
        
        det_headers = ['Task_ID', 'Task_Name', 'Task_Type', 'CT (초)', 'Station', 'Operator', 'Start (초)', 'End (초)']
        det_widths = [12, 30, 10, 10, 12, 14, 10, 10]
        
        det_header_fill = PatternFill('solid', start_color='305496')
        det_header_font = Font(name='Malgun Gothic', size=11, bold=True, color='FFFFFF')

        for c_idx, (h, w) in enumerate(zip(det_headers, det_widths), 1):
            cell = ws_det.cell(row=3, column=c_idx, value=h)
            cell.fill = det_header_fill
            cell.font = det_header_font
            cell.alignment = center_align
            col_letter = openpyxl.utils.get_column_letter(c_idx)
            ws_det.column_dimensions[col_letter].width = w
        ws_det.row_dimensions[3].height = 22

        # 할당 태스크들을 Station -> Operator -> Start 순서로 정렬
        assignment = res['best_result']['assignment']
        rows = []
        for tid, info in assignment.items():
            t = tasks[tid]
            rows.append({
                'task_id': tid,
                'task_name': t['task_name'],
                'task_type': t['task_type'],
                'ct': info['ct'],
                'station': info['station'],
                'operator': info['operator_label'],
                'start': info['start'],
                'end': info['end']
            })
        rows.sort(key=lambda r: (r['station'], r['operator'], r['start']))

        # 데이터 입력
        for r_idx, r in enumerate(rows, 4):
            vals = [
                r['task_id'], r['task_name'], r['task_type'], r['ct'],
                f"St-{r['station']:02d}", r['operator'], r['start'], r['end']
            ]
            for c_idx, v in enumerate(vals, 1):
                cell = ws_det.cell(row=r_idx, column=c_idx, value=v)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border
                cell.alignment = center_align if c_idx not in (2,) else left_align
                
                if c_idx in [4, 7, 8]:
                    cell.number_format = '0.0'

            # 가독성을 높이기 위해 줄 바꿈 색상 대비 추가
            if r_idx % 2 == 0:
                for c_idx in range(1, len(vals) + 1):
                    ws_det.cell(row=r_idx, column=c_idx).fill = PatternFill('solid', start_color='F2F2F2')
            ws_det.row_dimensions[r_idx].height = 18

    wb.save(report_path)
    print(f"[2/3] [Excel 완료] 비교 분석 리포트 저장 성공: {report_path}")


def plot_comparison_graphs(results, output_dir):
    """
    수집된 비교 데이터를 Matplotlib으로 시각화하여 고품질 PNG 파일로 저장합니다.
    """
    chart_path = os.path.join(output_dir, "comparison_chart.png")
    
    ns_list = [r['ns'] for r in results]
    ct_list = [r['cycle_time'] for r in results]
    eff_list = [r['efficiency'] * 100 for r in results]  # % 단위로 환산
    humans_list = [r['humans'] for r in results]
    robots_list = [r['robots'] for r in results]
    
    # 2행 1열 구조의 차트 생성
    fig, (ax1, ax3) = plt.subplots(2, 1, figsize=(11, 9), dpi=150)
    
    # ----------------------------------------------------
    # Subplot 1: Cycle Time & Line Efficiency vs NS (이중 Y축 적용)
    # ----------------------------------------------------
    color = '#1f77b4'  # Deep Blue
    ax1.set_xlabel('Number of Stations (NS)', fontweight='bold', labelpad=10)
    ax1.set_ylabel('Cycle Time (seconds)', color=color, fontweight='bold')
    line1 = ax1.plot(ns_list, ct_list, color=color, marker='o', linewidth=2.5, markersize=8, label='Cycle Time')
    ax1.tick_params(axis='y', labelcolor=color)
    ax1.grid(True, linestyle='--', alpha=0.5)
    
    # 우측 Y축 (라인 효율) 생성
    ax2 = ax1.twinx()  
    color = '#2ca02c'  # Forest Green
    ax2.set_ylabel('Line Efficiency (%)', color=color, fontweight='bold')
    line2 = ax2.plot(ns_list, eff_list, color=color, marker='s', linewidth=2.5, markersize=8, label='Line Efficiency')
    ax2.tick_params(axis='y', labelcolor=color)
    
    # 범례 설정
    lines = line1 + line2
    labels = [l.get_label() for l in lines]
    ax1.legend(lines, labels, loc='upper center', bbox_to_anchor=(0.5, 1.15), ncol=2, frameon=True, shadow=False)
    ax1.set_title('Cycle Time & Line Efficiency Trends by Station Count', pad=25, fontweight='bold', fontsize=12)
    ax1.set_xticks(ns_list)
    
    # ----------------------------------------------------
    # Subplot 2: Worker Count (Humans vs Robots) vs NS (그룹 바차트)
    # ----------------------------------------------------
    x = np.array(ns_list)
    width = 0.35
    
    rects1 = ax3.bar(x - width/2, humans_list, width, label='Humans (NH)', color='#ff7f0e')  # Orange
    rects2 = ax3.bar(x + width/2, robots_list, width, label='Robots (NR)', color='#1f77b4')  # Blue
    
    ax3.set_xlabel('Number of Stations (NS)', fontweight='bold', labelpad=10)
    ax3.set_ylabel('Number of Active Operators', fontweight='bold')
    ax3.set_title('Active Workers & Robots Distribution by Station Count', pad=15, fontweight='bold', fontsize=12)
    ax3.set_xticks(x)
    ax3.legend(loc='upper right')
    ax3.grid(True, axis='y', linestyle='--', alpha=0.5)
    
    # 막대 기둥마다 숫자 라벨링 추가
    def autolabel(rects):
        for rect in rects:
            height = rect.get_height()
            ax3.annotate(f'{int(height)}',
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),  # 3포인트 위로 오프셋
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=9)
            
    autolabel(rects1)
    autolabel(rects2)
    
    plt.tight_layout()
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"[3/3] [Chart 완료] 비교 분석 시각화 차트 저장 성공: {chart_path}")


def main():
    parser = argparse.ArgumentParser(
        description='SP3 CTR FLR 스테이션 수(10~20개)별 최적화 공정 비교 및 결과 정리 스크립트',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('input_file', help='입력 엑셀 파일 경로')
    parser.add_argument('--mode', choices=['quick', 'balanced', 'standard'], default='balanced',
                        help='실행 모드 (quick: 초고속, balanced: 권장 균형, standard: 정밀)')
    parser.add_argument('--output-dir', default='comparison_results',
                        help='결과 파일들을 생성하고 저장할 폴더명 (기본: comparison_results)')
    parser.add_argument('--seed', type=int, default=42,
                        help='랜덤 시드 고정 값 (기본: 42)')
    args = parser.parse_args()

    # 결과 폴더 생성
    os.makedirs(args.output_dir, exist_ok=True)

    # 1. 10~20 스테이션 루프 가동 및 연산 수집
    t_start_all = time.time()
    results, tasks = run_comparison(args.input_file, mode=args.mode, seed=args.seed)
    
    if not results:
        print("\n[ERROR] 수집된 최적화 결과가 없어 보고서를 작성할 수 없습니다.")
        sys.exit(1)

    # 2. 결과들을 취합하여 단일 엑셀 파일 저장
    write_comparison_excel(results, tasks, args.output_dir)

    # 3. 사이클 타임 및 인력 분포 시각화 차트 그리기
    plot_comparison_graphs(results, args.output_dir)

    elapsed_all = time.time() - t_start_all
    print(f"\n{'='*70}")
    print(f"  [SUCCESS] 모든 작업이 성공적으로 완료되었습니다! (총 소요시간: {elapsed_all:.1f}초)")
    print(f"  [Folder] 생성된 결과물 폴더: {os.path.abspath(args.output_dir)}")
    print(f"  - [Excel] 비교 엑셀 리포트: comparison_report.xlsx")
    print(f"  - [Chart] 비교 시각화 그래프: comparison_chart.png")
    print(f"{'='*70}\n")


if __name__ == '__main__':
    main()
