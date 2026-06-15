"""
엑셀 입출력 모듈
입력 스키마: BOM_Data | Task_Master | Precedence
출력: 결과 + 간트 차트 데이터
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def read_input(filepath):
    """
    엑셀 입력 읽기.

    Task_Master는 CT 컬럼 1개만 있는 단순 버전 / 또는 CT_human, CT_robot으로
    분리된 버전 모두 지원. 단순 버전이면 Task_Type에 따라 자동 분기:
      - Type 0 (인간전용): ct_human=CT, ct_robot=무한대
      - Type 1 (로봇전용): ct_human=무한대, ct_robot=CT
      - Type 2 (둘다가능): ct_human=ct_robot=CT  (로봇이 약간 빠르다고 가정해도 됨)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    required = ['BOM_Data', 'Task_Master', 'Precedence']
    for s in required:
        if s not in wb.sheetnames:
            raise ValueError(f"필수 시트 누락: '{s}' (시트 목록: {wb.sheetnames})")

    # ---- BOM_Data ----
    bom = _read_bom(wb['BOM_Data'])

    # ---- Task_Master ----
    tasks = _read_tasks(wb['Task_Master'])

    # ---- Precedence ----
    edges = _read_precedence(wb['Precedence'])

    return {'bom': bom, 'tasks': tasks, 'edges': edges}


def _read_bom(ws):
    """BOM_Data 시트 읽기"""
    headers = _read_headers(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        pn = _cell(ws, r, headers.get('Part_No'))
        if pn is None:
            continue
        rows.append({
            'level': _cell(ws, r, headers.get('Level')) or 0,
            'parent_pn': _cell(ws, r, headers.get('Parent_Part_No')),
            'part_no': pn,
            'part_name': _cell(ws, r, headers.get('Part_Name')) or '',
            'qty': _cell(ws, r, headers.get('Qty')) or 1,
            'variant': _cell(ws, r, headers.get('Variant')) or 'ALL',
        })
    return rows


def _read_tasks(ws):
    """Task_Master 시트 읽기 (안내 행 자동 skip)"""
    headers = _read_headers(ws)
    tasks = {}
    for r in range(2, ws.max_row + 1):
        tid = _cell(ws, r, headers.get('Task_ID'))
        if not tid or not isinstance(tid, str):
            continue
        # 안내 행 skip
        if tid.startswith('📌') or tid.startswith('#'):
            continue

        task_type = _safe_int(_cell(ws, r, headers.get('Task_Type')), 2)
        ct = _safe_float(_cell(ws, r, headers.get('CT')), 0.0)

        # Task_Type에 따라 CT_human / CT_robot 자동 결정
        # (분리 컬럼이 있으면 그것 우선)
        ct_h_col = headers.get('CT_Human') or headers.get('CT_human')
        ct_r_col = headers.get('CT_Robot') or headers.get('CT_robot')

        if ct_h_col and ct_r_col:
            ct_human = _safe_float(_cell(ws, r, ct_h_col), float('inf'))
            ct_robot = _safe_float(_cell(ws, r, ct_r_col), float('inf'))
        else:
            # CT 단일 컬럼 — Task_Type 기반 분기
            if task_type == 0:    # 인간전용
                ct_human, ct_robot = ct, float('inf')
            elif task_type == 1:  # 로봇전용
                ct_human, ct_robot = float('inf'), ct
            else:                 # 2 = 둘다가능
                ct_human, ct_robot = ct, ct

        tasks[tid] = {
            'task_id': tid,
            'task_name': _cell(ws, r, headers.get('Task_Name')) or '',
            'task_type': task_type,
            'ct_human': ct_human,
            'ct_robot': ct_robot,
            'output_pn': _cell(ws, r, headers.get('Output_Part_No')),
            'variant': _cell(ws, r, headers.get('Variant')) or 'ALL',
        }
    return tasks


def _read_precedence(ws):
    """Precedence 시트 읽기"""
    headers = _read_headers(ws)
    edges = []
    pred_col = headers.get('Predecessor_ID') or headers.get('Predecessor')
    succ_col = headers.get('Successor_ID') or headers.get('Successor')
    if pred_col is None or succ_col is None:
        raise ValueError("Precedence 시트에 Predecessor_ID / Successor_ID 컬럼 필요")
    for r in range(2, ws.max_row + 1):
        p = _cell(ws, r, pred_col)
        s = _cell(ws, r, succ_col)
        if p and s:
            edges.append((str(p).strip(), str(s).strip()))
    return edges


def _read_headers(ws):
    """헤더 행에서 컬럼명 → 컬럼 번호 매핑"""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip()] = c
    return headers


def _cell(ws, row, col):
    """안전한 셀 읽기"""
    if col is None:
        return None
    return ws.cell(row=row, column=col).value


def _safe_int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _safe_float(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


# ============================================================
# 결과 출력
# ============================================================
def write_result(filepath, data, best, args):
    """결과 엑셀 작성"""
    wb = Workbook()

    _write_summary_sheet(wb, best, args)
    _write_schedule_sheet(wb, best, data['tasks'])
    _write_gantt_sheet(wb, best, data['tasks'])
    _write_metrics_sheet(wb, best)

    # 첫 시트(active) 제거
    if wb.sheetnames[0] == 'Sheet':
        del wb['Sheet']

    wb.save(filepath)


def _write_summary_sheet(wb, best, args):
    """1. 요약 시트"""
    ws = wb.create_sheet('Summary')
    ws['B2'] = '자동 시퀀싱 결과 요약'
    ws['B2'].font = Font(name='Calibri', size=14, bold=True)

    rows = [
        ('', ''),
        ('입력 설정', ''),
        ('  스테이션 수 (NS)', args.stations),
        ('  스테이션당 최대 인간', args.humans),
        ('  스테이션당 최대 로봇', args.robots),
        ('', ''),
        ('최적화 결과', ''),
        ('  사이클타임 (Cycle Time)', f"{best['cycle_time']:.1f} 초"),
        ('  사용된 인간 수 (NH)', best['n_humans']),
        ('  사용된 로봇 수 (NR)', best['n_robots']),
        ('  Smoothness Index (SX)', f"{best['smoothness']:.2f}"),
        ('  병목 station', f"Station {best['bottleneck_station']} ({best['bottleneck_ct']:.1f}초)"),
        ('  라인 밸런스 효율', f"{best['line_efficiency']*100:.1f}%"),
        ('', ''),
        ('알고리즘', ''),
        ('  방법', 'Enhanced Adaptive Simulated Annealing (E_ASA)'),
        ('  베이스 논문', 'Nourmohammadi et al. (2022) + Zhang & Fujimura (2025)'),
        ('  실행 모드', 'Quick' if args.quick else 'Standard'),
        ('  반복 횟수', best.get('total_iterations', '-')),
    ]
    for i, (k, v) in enumerate(rows, 4):
        ws.cell(row=i, column=2, value=k).font = Font(name='Calibri', size=11,
                                                       bold=not k.startswith(' '))
        ws.cell(row=i, column=3, value=v).font = Font(name='Calibri', size=11)
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 50


def _write_schedule_sheet(wb, best, tasks):
    """2. 공정 할당 시트 (각 task의 station/operator/start/end)"""
    ws = wb.create_sheet('Schedule')
    headers = ['Task_ID', 'Task_Name', 'Type', 'CT', 'Station',
               'Operator', 'Start', 'End', 'Output_Part']
    _write_header(ws, headers, [12, 28, 8, 8, 10, 14, 10, 10, 22])

    # station 순서 → operator 순서 → start 순서로 정렬
    rows = []
    for tid, info in best['assignment'].items():
        t = tasks[tid]
        rows.append({
            'task_id': tid,
            'task_name': t['task_name'],
            'task_type': t['task_type'],
            'ct': info['ct'],
            'station': info['station'],
            'operator': info['operator_label'],
            'start': info['start'],
            'end': info['end'],
            'output_pn': t.get('output_pn', ''),
        })
    rows.sort(key=lambda r: (r['station'], r['operator'], r['start']))

    for i, r in enumerate(rows, 2):
        vals = [r['task_id'], r['task_name'], r['task_type'], r['ct'],
                f"St-{r['station']:02d}", r['operator'],
                f"{r['start']:.1f}", f"{r['end']:.1f}", r['output_pn'] or '']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='center' if c not in (2, 9) else 'left')
            if i % 2 == 0:
                cell.fill = PatternFill('solid', start_color='F2F2F2')
    ws.freeze_panes = 'A2'


def _write_gantt_sheet(wb, best, tasks):
    """3. Gantt 데이터 시트 (시각화용 raw data)"""
    ws = wb.create_sheet('Gantt_Data')
    headers = ['Station', 'Operator', 'Task_ID', 'Task_Name',
               'Start', 'End', 'Duration', 'Type']
    _write_header(ws, headers, [10, 14, 12, 28, 10, 10, 10, 8])

    rows = []
    for tid, info in best['assignment'].items():
        t = tasks[tid]
        rows.append([
            info['station'], info['operator_label'], tid, t['task_name'],
            info['start'], info['end'], info['ct'], t['task_type'],
        ])
    rows.sort(key=lambda r: (r[0], r[1], r[4]))

    for i, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = Font(name='Calibri', size=10)
    ws.freeze_panes = 'A2'


def _write_metrics_sheet(wb, best):
    """4. Station별 메트릭"""
    ws = wb.create_sheet('Station_Metrics')
    headers = ['Station', 'Operators', 'Load (sec)', 'Utilization (%)',
               'Idle (sec)', 'Tasks_Count']
    _write_header(ws, headers, [10, 18, 14, 16, 12, 14])

    for i, s in enumerate(best['station_metrics'], 2):
        vals = [f"St-{s['station']:02d}", s['operators'],
                f"{s['load']:.1f}", f"{s['utilization']*100:.1f}",
                f"{s['idle']:.1f}", s['n_tasks']]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'


def _write_header(ws, headers, widths):
    """공통 헤더 작성"""
    fill = PatternFill('solid', start_color='305496')
    font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22


def print_summary(best, tasks):
    """터미널 출력용 요약"""
    print(f"\n  사이클타임 (CT):     {best['cycle_time']:.1f} 초")
    print(f"  사용 인간 수 (NH):   {best['n_humans']}")
    print(f"  사용 로봇 수 (NR):   {best['n_robots']}")
    print(f"  Smoothness Index:    {best['smoothness']:.2f}")
    print(f"  병목 Station:        St-{best['bottleneck_station']:02d} ({best['bottleneck_ct']:.1f}초)")
    print(f"  라인 밸런스 효율:    {best['line_efficiency']*100:.1f}%")

    print(f"\n  Station별 부하:")
    for s in best['station_metrics']:
        bar_len = int(s['utilization'] * 30)
        bar = '#' * bar_len + '-' * (30 - bar_len)
        marker = ' <= 병목' if s['station'] == best['bottleneck_station'] else ''
        print(f"    St-{s['station']:02d}  [{bar}]  {s['utilization']*100:5.1f}%  "
              f"({s['load']:5.1f}s, {s['n_tasks']:2d}개){marker}")
